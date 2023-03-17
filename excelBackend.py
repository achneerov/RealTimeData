
from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import workbook

workbook = load_workbook("RealTimeData.xlsx")
sheet = workbook.active
sheet["C1"] = "writing ;)"
workbook.save("RealTimeData.xlsx")

#data = pd.read_excel("RealTimeData.xlsx")

print(sheet["A2"])
print(sheet.cell(row=2, column=2).value)





# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])



#y = data['age'][:50]
#x = []
#for a in range(1, 51):
    #x.append(a)
#plt.plot(x,y)
#plt.show()